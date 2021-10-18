import xlwt
import time,os,glob
import openpyxl
from openpyxl.cell.cell import get_column_letter
from  openpyxl  import Workbook
from sendemail import Email_li
import configparser
cf = configparser.ConfigParser()
cf.read("./conf.ini", encoding="utf-8")


class exccell_li:
    #def __init__(self,fields_complex,title):
    def __init__(self, **kwargs):
        self.title = kwargs['title']
        try:
            self.fields = []  # 要用到中文字段名 就需要注释
            for i in range(len(kwargs['fields_complex'])):
                self.fields.append(kwargs['fields_complex'][i][0])
        except:
            print("不需要自定义字段")
    def  write_to_excel_with_openpyxl(self,all_data):
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = self.title

        data = []
        try:
            data.append(self.fields)
        except:
            print("不需要自定义列名")
        for i in all_data:
            data.append(i)
        i = 1
        for line in data:
            for col in range(1, len(line) + 1):
                ColNum = get_column_letter(col)
                #ws.cell(row='%s', column='%s' %(i,col)).value='%s' % (line[col - 1])
                ws.cell(row=i, column=col, value="{0}".format(line[col - 1]))
            i += 1
        wb.save('{0}/{1}{2}.xlsx'.format(cf.get('excell','out_file_dir'),self.title,str(time.strftime('%Y-%m-%d'))))
        print("导出文件成功")

    def send_email(self):
        #读取配置文件，拼接和获取需要发送邮件目录下的文件路径
        file_path=os.path.join(os.getcwd(),cf.get('excell','out_file_dir'))
        #print(glob.glob(file_path+ r'\*.xlsx'))
        file_lists=[]
        for i in os.listdir(file_path):
            file_lists.append(os.path.join(file_path,i))
        #print(file_lists)
        #cf.set('email','reports_path',str(file_lists))
        #cf.write(open("./conf.ini","w",encoding="utf=-8"))
        #开始发送邮件
        s_email=Email_li()
        # 发送多个附件，需要列表的形式，这里直接传参过去
        s_email.send(reports_path=file_lists)





        #self.save(filename=r'C:\Users\Administrator\Downloads\yc20210322\{}.xls'.format(table_name))
